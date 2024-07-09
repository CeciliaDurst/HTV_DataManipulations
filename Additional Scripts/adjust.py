import openpyxl


def adjust_excel(input_file, output_file):
    wb = openpyxl.load_workbook(input_file)

    # Iterate through each sheet in the workbook
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        for col in range(1, ws.max_column + 1):
            empty_cells_count = 0
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is None:
                    empty_cells_count += 1
                else:
                    break

            if empty_cells_count > 0:
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col)
                    new_row = row - empty_cells_count
                    if new_row > 1:
                        new_cell = ws.cell(row=new_row, column=col)
                        new_cell.value = cell.value
                        cell.value = None

    wb.save(output_file)


# Example usage
input_file = "test200.xlsx"
output_file = "test201.xlsx"
adjust_excel(input_file, output_file)
