import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sklearn.impute import KNNImputer

file_path = "F23_rawSleepEfficiency_shifted_filtered.xlsx"
start_row = 2
df = pd.read_excel(file_path, header=None, skiprows=start_row - 1)

# Replace empty strings with NaN
df.replace("", pd.NA, inplace=True)

imputer = KNNImputer(n_neighbors=5)

# Interpolate missing values using KNNImputer for each column separately
for col in df.columns:
    # Create a DataFrame with only the current column
    col_df = pd.DataFrame(df[col])

    # Impute missing values for the current column
    col_df_imputed = pd.DataFrame(imputer.fit_transform(col_df), columns=[col])

    # Update the original DataFrame with imputed values for the current column
    df[col] = col_df_imputed[col]

output_file = "F23_rawSleepEfficiency_imputedz11.xlsx"
df.to_excel(output_file, index=False, header=False)

wb = load_workbook(output_file)
ws = wb.active
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

# Apply conditional formatting to highlight imputed cells for each column separately
for index, row in df.iterrows():
    for col in df.columns:
        if pd.isna(row[col]):
            row_idx = index + 1
            ws.cell(row=row_idx, column=col + 1).fill = yellow_fill  # +1 to adjust for 1-based indexing

wb.save(output_file)
