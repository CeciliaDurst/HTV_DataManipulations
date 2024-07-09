import csv
from datetime import datetime
from openpyxl import Workbook
import data as data
import data
count_2_minutes_or_less = 0

def convert_to_datetime(time_str):
    try:
        return datetime.strptime(time_str, '%m/%d/%Y %I:%M:%S %p')
    except ValueError:
        try:
            return datetime.strptime(time_str, '%m/%d/%Y %I:%M %p')
        except ValueError:
            print("Error: Unable to parse time data:", time_str)
            return None

def calculate_sleep_latency(entries):
    wake_times = []
    light_sleep_time = None

    for entry in entries:
        time = convert_to_datetime(entry[1])
        level = entry[3]

        if level == 'wake':
            if 18 <= time.hour < 24:
                wake_times.append(time)
        elif level == 'light' or level == 'rem':
            if time.hour >= 18:
                light_sleep_time = time
                break

    if wake_times and light_sleep_time:
        wake_times_before_light_sleep = [wake for wake in wake_times if wake < light_sleep_time]

        if len(wake_times_before_light_sleep) > 4:
            first_wake_after_6pm = wake_times_before_light_sleep[0]
            last_wake_before_light_sleep = wake_times_before_light_sleep[-1]
            sleep_latency = (last_wake_before_light_sleep - first_wake_after_6pm).total_seconds() / 60

            print(
                f"First Wake after 6 PM: {first_wake_after_6pm}, Last Wake before 'light' sleep: {last_wake_before_light_sleep}, Sleep Latency: {sleep_latency} minutes")
            if sleep_latency <= 2:
                global count_2_minutes_or_less
                count_2_minutes_or_less += 1
            return sleep_latency, first_wake_after_6pm, last_wake_before_light_sleep

    return None, None, None


# Read data from CSV file
file_name = 'SP23_30SecondSleepStages.csv'
with open(file_name, 'r') as file:
    reader = csv.reader(file)
    data = list(reader)
# Create Excel workbook and sheet
wb = Workbook()
ws = wb.active

# Write headers
ws['A1'] = 'Participant ID'
ws['B1'] = 'Date'
ws['C1'] = 'First Wake before light sleep'
ws['D1'] = 'Last Wake before light sleep'
ws['E1'] = 'Sleep latency'

# Group data by participant ID and date
grouped_data = {}
for entry in data[1:]:  # Skip header row
    participant_id = entry[0]
    date = entry[1].split()[0]  # Extract date portion of time
    key = (participant_id, date)

    if key not in grouped_data:
        grouped_data[key] = []
    grouped_data[key].append(entry)

row = 2
for key, entries in grouped_data.items():
    participant_id, date = key
    sleep_latency, first_wake, last_wake = calculate_sleep_latency(entries)
    ws.cell(row=row, column=1, value=participant_id)
    ws.cell(row=row, column=2, value=date)
    if sleep_latency is not None:
        ws.cell(row=row, column=3, value=first_wake)
        ws.cell(row=row, column=4, value=last_wake)
        # Append 'minutes' to sleep latency value
        ws.cell(row=row, column=5, value=sleep_latency)
    else:
        ws.cell(row=row, column=3, value=0)
        ws.cell(row=row, column=4, value=0)
        ws.cell(row=row, column=5, value=0)
    row += 1

# Adjust column widths
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length
# Save Excel file
wb.save('SP2023sleep_latency_results.xlsx')